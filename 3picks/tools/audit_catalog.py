#!/usr/bin/env python3
"""상품마스터와 브라우저용 상품 데이터의 핵심 필드를 전수 대조합니다."""

from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
MASTER = ROOT.parent / "굿즈리스트 전달용" / "[최종본] 상품마스터.xlsx"
RUNTIME = ROOT / "assets" / "products-data.js"


def load_runtime() -> list[dict]:
    source = RUNTIME.read_text(encoding="utf-8").strip()
    prefix = "window.PRODUCTS="
    if not source.startswith(prefix) or not source.endswith(";"):
        raise AssertionError("products-data.js 형식이 예상과 다릅니다.")
    return json.loads(source[len(prefix) : -1])


def load_master() -> list[dict]:
    workbook = load_workbook(MASTER, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    return [dict(zip(headers, row)) for row in rows if row[0] is not None]


master = load_master()
runtime = load_runtime()
assert len(master) == 110, f"마스터 행 수 오류: {len(master)}"
assert len(runtime) == 110, f"런타임 행 수 오류: {len(runtime)}"

master_by_number = {int(row["No"]): row for row in master}
runtime_by_number = {int(row["number"]): row for row in runtime}
assert len(master_by_number) == len(runtime_by_number) == 110, "중복 번호가 있습니다."
assert set(master_by_number) == set(runtime_by_number), "마스터와 런타임의 상품 번호가 다릅니다."

mismatches: list[str] = []
for number, source in master_by_number.items():
    product = runtime_by_number[number]
    checks = {
        "category": (re.sub(r"^\d+\.\s*", "", str(source["카테고리"])), product["category"]),
        "name": (source["상품명"], product["name"]),
        "visibility": (source["★노출구분"], product["visibility"]),
        "price": (int(source["객단가(100개,원)"]), int(product["price"])),
    }
    for field, (left, right) in checks.items():
        if left != right:
            mismatches.append(f"No.{number} {field}: {left!r} != {right!r}")

missing_images = [
    f"{product['id']}:{image}"
    for product in runtime
    for image in product.get("images", [])
    if not (ROOT / image).is_file()
]
invalid_image_counts = [product["id"] for product in runtime if len(product.get("images", [])) != 3]
duplicate_names = [name for name, count in Counter(product["name"] for product in runtime).items() if count > 1]
eligible = [
    product
    for product in runtime
    if product["visibility"] == "화면노출"
    and product["available"]
    and product.get("price")
    and product.get("moq")
    and product.get("leadDays")
]

category_prices: dict[str, list[int]] = defaultdict(list)
for product in runtime:
    category_prices[product["category"]].append(int(product["price"]))

assert not mismatches, "\n".join(mismatches[:20])
assert not missing_images, f"이미지 파일 누락: {missing_images[:10]}"
assert not invalid_image_counts, f"3색 이미지 수 오류: {invalid_image_counts}"
assert not duplicate_names, f"중복 상품명: {duplicate_names}"

report = {
    "master_rows": len(master),
    "runtime_rows": len(runtime),
    "core_field_mismatches": len(mismatches),
    "priced": sum(bool(product.get("price")) for product in runtime),
    "available": sum(bool(product.get("available")) for product in runtime),
    "screen_visible": sum(product.get("visibility") == "화면노출" for product in runtime),
    "recommendation_eligible": len(eligible),
    "three_images": sum(len(product.get("images", [])) == 3 for product in runtime),
    "missing_image_files": len(missing_images),
    "category_price_ranges": {
        category: {"count": len(prices), "min": min(prices), "max": max(prices)}
        for category, prices in category_prices.items()
    },
}
print(json.dumps(report, ensure_ascii=False, indent=2))
